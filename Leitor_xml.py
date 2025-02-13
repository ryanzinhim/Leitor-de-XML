import os
import glob
import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# Configuração do logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

DOWNLOAD_PATH = "C:\\Users\\ryan.alves\\Downloads\\XML"
EXCEL_PATH = "C:\\Users\\ryan.alves\\Downloads\\Sem título 1.xlsx"
SHEET_NAME = "Número e chave de acesso - Mone"

def load_excel():
    """Carrega a planilha e retorna a aba de trabalho."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    ws['A1'] = "Numero"
    ws['B1'] = "Chave"
    return wb, ws

def process_xml(xml_file):
    """Processa um único arquivo XML e retorna número da NF e chave de acesso."""
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
        
        nNF = root.find(".//nfe:nNF", ns)
        chave = root.find(".//nfe:infNFe", ns)
        
        nNF = int(nNF.text) if nNF is not None else "ERRO"
        chave = chave.attrib.get('Id', "")[3:] if chave is not None else "ERRO"

        return nNF, chave
    except Exception as e:
        logging.error(f"Erro ao processar {xml_file}: {e}")
        return "ERRO", "ERRO"

def process_xmls():
    """Processa todos os arquivos XML no diretório e salva no Excel."""
    try:
        logging.info("Iniciando processamento dos XMLs...")
        wb, ws = load_excel()
        xml_files = glob.glob(os.path.join(DOWNLOAD_PATH, "*.xml"))
        total_files = len(xml_files)
        found_numbers = set()
        duplicate_numbers = set()  # Para controlar os números duplicados
        row = 2
        missing_numbers = set()  # Para controlar números faltantes

        for i, xml_file in enumerate(xml_files, start=1):
            nNF, chave = process_xml(xml_file)

            if nNF != "ERRO":
                if nNF in found_numbers:
                    duplicate_numbers.add(nNF)
                    logging.warning(f"Arquivo repetido: {xml_file}")
                found_numbers.add(nNF)
            
            ws[f"A{row}"] = str(nNF)
            ws[f"B{row}"] = chave
            row += 1
            logging.info(f"Processado {i}/{total_files}: NF {nNF}, Chave {chave}")

        ##Criação do intervalo de NFs esperadas de 22393 até 30664
        expected_numbers = set(range(22393, 30665))
        missing_numbers = expected_numbers - found_numbers
        if missing_numbers:
            logging.warning(f"Faltando NFs: {', '.join(map(str, missing_numbers))}")

        ##Agora, removemos as duplicatas antes de salvar
        if duplicate_numbers:
            logging.info(f"Removendo duplicatas: {', '.join(map(str, duplicate_numbers))}")
            ws.delete_rows(2, row - 1)  # Apaga as linhas existentes
            row = 2  # Redefine a contagem da linha
            for nNF, chave in zip(found_numbers - duplicate_numbers, [ws[f"B{r}"].value for r in range(2, row)]):
                ws[f"A{row}"] = str(nNF)
                ws[f"B{row}"] = chave
                row += 1
               
            return len(found_numbers - duplicate_numbers)

        wb.save(EXCEL_PATH)
        logging.info("Processamento concluído!")
    except Exception as e:
        logging.error(f"Erro no processamento dos XMLs: {e}")

if __name__ == "__main__":
    process_xmls()
